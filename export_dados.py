#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Script para exportar dados de avaliações
Uso:
    python export_dados.py excel                          # Exporta tudo para Excel
    python export_dados.py txt                            # Exporta tudo para TXT
    python export_dados.py txt --inicio 2026-01-01 --fim 2026-01-31  # Exporta período
"""

import sqlite3
import csv
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

DATABASE = 'satisfaction.db'

def conectar_db():
    """Conecta ao banco de dados"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def exportar_excel(nome_arquivo=None):
    """Exporta dados para Excel (.xlsx)"""
    if nome_arquivo is None:
        nome_arquivo = f'avaliacoes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    try:
        conn = conectar_db()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, grau_satisfacao, data, hora, dia_semana 
            FROM avaliacoes 
            ORDER BY id ASC
        ''')
        rows = cursor.fetchall()
        conn.close()
        
        if not rows:
            print("❌ Nenhum dado para exportar!")
            return
        
        if HAS_OPENPYXL:
            # Criar arquivo XLSX com formatação
            wb = Workbook()
            ws = wb.active
            ws.title = "Avaliações"
            
            # Estilos
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Cabeçalho
            headers = ['ID', 'Grau de Satisfação', 'Data', 'Hora', 'Dia da Semana']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Dados
            emoji_map = {
                'muito_satisfeito': '😊 Muito Satisfeito',
                'satisfeito': '😐 Satisfeito',
                'insatisfeito': '😞 Insatisfeito'
            }
            
            for row_idx, row in enumerate(rows, start=2):
                satisfacao_display = emoji_map.get(row[1], row[1])
                
                data = [row[0], satisfacao_display, row[2], row[3], row[4]]
                for col, value in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.value = value
                    cell.border = border
                    if col == 1 or col == 3 or col == 4:
                        cell.alignment = Alignment(horizontal='center')
            
            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 18
            
            wb.save(nome_arquivo)
            print(f"✅ Exportado para Excel: {nome_arquivo}")
            print(f"📊 Total de registros: {len(rows)}")
        
        else:
            # Fallback para CSV se openpyxl não está instalado
            print("⚠️  openpyxl não instalado. Exportando para CSV...")
            nome_arquivo = nome_arquivo.replace('.xlsx', '.csv')
            
            with open(nome_arquivo, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f, delimiter=';', quoting=csv.QUOTE_ALL)
                writer.writerow(['ID', 'Grau de Satisfação', 'Data', 'Hora', 'Dia da Semana'])
                for row in rows:
                    writer.writerow([row[0], row[1], row[2], row[3], row[4]])
            
            print(f"✅ Exportado para CSV: {nome_arquivo}")
            print(f"📊 Total de registros: {len(rows)}")
        
    except Exception as e:
        print(f"❌ Erro ao exportar para Excel: {e}")
        import traceback
        traceback.print_exc()

def exportar_txt(data_inicio=None, data_fim=None, nome_arquivo=None):
    """Exporta dados para TXT com filtro opcional"""
    if nome_arquivo is None:
        nome_arquivo = f'avaliacoes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'
    
    try:
        conn = conectar_db()
        cursor = conn.cursor()
        
        if data_inicio and data_fim:
            cursor.execute('''
                SELECT id, grau_satisfacao, data, hora, dia_semana 
                FROM avaliacoes 
                WHERE data BETWEEN ? AND ?
                ORDER BY id ASC
            ''', (data_inicio, data_fim))
        else:
            cursor.execute('''
                SELECT id, grau_satisfacao, data, hora, dia_semana 
                FROM avaliacoes 
                ORDER BY id ASC
            ''')
        
        rows = cursor.fetchall()
        conn.close()
        
        if not rows:
            print("❌ Nenhum dado para exportar!")
            return
        
        with open(nome_arquivo, 'w', encoding='utf-8') as f:
            f.write('=' * 100 + '\n')
            f.write('RELATÓRIO DE AVALIAÇÕES DE SATISFAÇÃO\n')
            f.write('=' * 100 + '\n\n')
            
            if data_inicio and data_fim:
                f.write(f'Período: {data_inicio} até {data_fim}\n')
            else:
                f.write(f'Data de Geração: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}\n')
            
            f.write(f'Total de Registros: {len(rows)}\n')
            f.write('=' * 100 + '\n\n')
            
            # Tabela formatada
            f.write(f"{'ID':<6} | {'Satisfação':<25} | {'Data':<12} | {'Hora':<10} | {'Dia da Semana':<15}\n")
            f.write('-' * 100 + '\n')
            
            for row in rows:
                emoji = '😊' if row[1] == 'muito_satisfeito' else '😐' if row[1] == 'satisfeito' else '😞'
                satisfacao_texto = f"{emoji} {row[1].replace('_', ' ').title()}"
                
                f.write(f"{row[0]:<6} | {satisfacao_texto:<25} | {row[2]:<12} | {row[3]:<10} | {row[4]:<15}\n")
            
            f.write('\n' + '=' * 100 + '\n')
            f.write(f'Fim do Relatório\n')
            f.write('=' * 100 + '\n')
        
        print(f"✅ Exportado para TXT: {nome_arquivo}")
        print(f"📄 Total de registros: {len(rows)}")
        
    except Exception as e:
        print(f"❌ Erro ao exportar para TXT: {e}")
        import traceback
        traceback.print_exc()

def mostrar_ajuda():
    """Mostra mensagem de ajuda"""
    print("\n" + "=" * 80)
    print("📤 EXPORTADOR DE DADOS - AVALIAÇÕES DE SATISFAÇÃO")
    print("=" * 80 + "\n")
    print("Uso:")
    print("  python export_dados.py excel                              # Exporta tudo para Excel")
    print("  python export_dados.py txt                                # Exporta tudo para TXT")
    print("  python export_dados.py txt --inicio YYYY-MM-DD --fim YYYY-MM-DD")
    print("                                                             # Exporta período para TXT\n")
    print("Exemplos:")
    print("  python export_dados.py excel")
    print("  python export_dados.py txt --inicio 2026-01-01 --fim 2026-01-31\n")
    print("=" * 80 + "\n")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        mostrar_ajuda()
        sys.exit(1)
    
    comando = sys.argv[1].lower()
    
    if comando == 'excel':
        exportar_excel()
    
    elif comando == 'txt':
        data_inicio = None
        data_fim = None
        
        # Parse dos argumentos
        if '--inicio' in sys.argv:
            idx = sys.argv.index('--inicio')
            if idx + 1 < len(sys.argv):
                data_inicio = sys.argv[idx + 1]
        
        if '--fim' in sys.argv:
            idx = sys.argv.index('--fim')
            if idx + 1 < len(sys.argv):
                data_fim = sys.argv[idx + 1]
        
        exportar_txt(data_inicio, data_fim)
    
    else:
        print(f"❌ Comando desconhecido: {comando}")
        mostrar_ajuda()
        sys.exit(1)
