from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from flask_cors import CORS
import sqlite3
from datetime import datetime
import os
from contextlib import contextmanager
from io import StringIO, BytesIO
import csv
from functools import wraps

try:
    from pymongo import MongoClient
    HAS_PYMONGO = True
except ImportError:
    HAS_PYMONGO = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

app = Flask(__name__)
CORS(app)

# Configuração de sessão
app.secret_key = os.environ.get('SECRET_KEY', 'sua-chave-secreta-mude-em-producao')

# Código de acesso ao admin (pode ser alterado via variável de ambiente)
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', '1234')

# Decorator para proteger rotas admin
def login_required(f):
    """Verifica se o utilizador está autenticado"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_authenticated' not in session:
            return redirect(url_for('login_admin'))
        return f(*args, **kwargs)
    return decorated_function

# Configuração do banco de dados
DATABASE = 'satisfaction.db'
MONGODB_URI = os.environ.get('MONGODB_URI')
MONGODB_DB = os.environ.get('MONGODB_DB', 'testelp')

mongo_client = None
mongo_db = None
mongo_collection = None
USE_MONGO = False

if HAS_PYMONGO and MONGODB_URI:
    mongo_client = MongoClient(MONGODB_URI)
    mongo_db = mongo_client.get_default_database()
    if mongo_db is None:
        mongo_db = mongo_client[MONGODB_DB]
    mongo_collection = mongo_db['avaliacoes']
    USE_MONGO = True

@contextmanager
def get_db():
    """Context manager para conexão com o banco de dados"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()

def init_db():
    """Inicializa o banco de dados"""
    if USE_MONGO:
        return

    with get_db() as conn:
        conn.execute('''
            CREATE TABLE IF NOT EXISTS avaliacoes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                grau_satisfacao TEXT NOT NULL,
                data TEXT NOT NULL,
                hora TEXT NOT NULL,
                dia_semana TEXT NOT NULL,
                timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.commit()

# Inicializar banco de dados ao iniciar a aplicação
init_db()

def _mongo_filtro_data(data_inicio=None, data_fim=None):
    if data_inicio and data_fim:
        return {'data': {'$gte': data_inicio, '$lte': data_fim}}
    return {}

def _mongo_contar_por_campo(campo, data_inicio=None, data_fim=None):
    filtro = _mongo_filtro_data(data_inicio, data_fim)
    pipeline = [
        {'$match': filtro},
        {'$group': {'_id': f'${campo}', 'count': {'$sum': 1}}}
    ]
    return {row['_id']: row['count'] for row in mongo_collection.aggregate(pipeline)}

def _mongo_avaliacoes_diarias(data_inicio=None, data_fim=None):
    filtro = _mongo_filtro_data(data_inicio, data_fim)
    pipeline = [
        {'$match': filtro},
        {'$group': {'_id': '$data', 'count': {'$sum': 1}}},
        {'$sort': {'_id': 1}}
    ]
    return [{'data': row['_id'], 'count': row['count']} for row in mongo_collection.aggregate(pipeline)]

def _mongo_ultimas_avaliacoes(pagina, limite, data_inicio=None, data_fim=None):
    filtro = _mongo_filtro_data(data_inicio, data_fim)
    offset = (pagina - 1) * limite
    cursor = (mongo_collection.find(filtro)
              .sort('_id', -1)
              .skip(offset)
              .limit(limite))

    return [
        {
            'id': str(doc.get('_id')),
            'grau_satisfacao': doc.get('grau_satisfacao'),
            'data': doc.get('data'),
            'hora': doc.get('hora'),
            'dia_semana': doc.get('dia_semana')
        }
        for doc in cursor
    ]

def _mongo_listar_avaliacoes(data_inicio=None, data_fim=None, asc=True):
    filtro = _mongo_filtro_data(data_inicio, data_fim)
    ordem = 1 if asc else -1
    cursor = mongo_collection.find(filtro).sort('_id', ordem)
    return [
        (
            str(doc.get('_id')),
            doc.get('grau_satisfacao'),
            doc.get('data'),
            doc.get('hora'),
            doc.get('dia_semana')
        )
        for doc in cursor
    ]

@app.route('/')
def index():
    """Página principal de votação"""
    return render_template('index.html')

@app.route('/login-admin', methods=['GET', 'POST'])
def login_admin():
    """Página de login para o admin"""
    if request.method == 'POST':
        password = request.form.get('password', '')
        
        if password == ADMIN_PASSWORD:
            session['admin_authenticated'] = True
            return redirect(url_for('admin'))
        else:
            return render_template('login.html', error='Código de acesso incorreto!')
    
    # Se já está autenticado, vai direto ao admin
    if 'admin_authenticated' in session:
        return redirect(url_for('admin'))
    
    return render_template('login.html')

@app.route('/admin')
@login_required
def admin():
    """Página administrativa"""
    return render_template('admin.html')

@app.route('/logout')
def logout():
    """Faz logout do admin"""
    session.pop('admin_authenticated', None)
    return redirect(url_for('index'))

@app.route('/api/vote', methods=['POST'])
def vote():
    """Registra um voto"""
    try:
        data = request.get_json()
        satisfacao = data.get('satisfacao')
        
        if satisfacao not in ['muito_satisfeito', 'satisfeito', 'insatisfeito']:
            return jsonify({'error': 'Opção inválida'}), 400
        
        # Obter data e hora atuais
        agora = datetime.now()
        data_str = agora.strftime('%Y-%m-%d')
        hora_str = agora.strftime('%H:%M:%S')
        
        # Dias da semana em português
        dias_semana = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']
        dia_semana = dias_semana[agora.weekday()]
        
        # Inserir no banco de dados
        if USE_MONGO:
            mongo_collection.insert_one({
                'grau_satisfacao': satisfacao,
                'data': data_str,
                'hora': hora_str,
                'dia_semana': dia_semana,
                'timestamp': datetime.utcnow()
            })
        else:
            with get_db() as conn:
                conn.execute(
                    'INSERT INTO avaliacoes (grau_satisfacao, data, hora, dia_semana) VALUES (?, ?, ?, ?)',
                    (satisfacao, data_str, hora_str, dia_semana)
                )
                conn.commit()
        
        return jsonify({
            'success': True,
            'message': 'Obrigado pelo seu feedback!'
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/stats', methods=['GET'])
def get_stats():
    """Retorna estatísticas das avaliações"""
    try:
        # Parâmetros de filtro
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        # Últimas avaliações (paginação)
        pagina = int(request.args.get('pagina', 1))
        limite = 20

        if USE_MONGO:
            total = mongo_collection.count_documents(_mongo_filtro_data(data_inicio, data_fim))
            satisfacao_dict = _mongo_contar_por_campo('grau_satisfacao', data_inicio, data_fim)
            dia_dict = _mongo_contar_por_campo('dia_semana', data_inicio, data_fim)
            avaliacoes_diarias = _mongo_avaliacoes_diarias(data_inicio, data_fim)
            ultimas_avaliacoes = _mongo_ultimas_avaliacoes(pagina, limite, data_inicio, data_fim)
            total_registos = total
        else:
            with get_db() as conn:
                # Total de avaliações
                if data_inicio and data_fim:
                    total = conn.execute(
                        'SELECT COUNT(*) as count FROM avaliacoes WHERE data BETWEEN ? AND ?',
                        (data_inicio, data_fim)
                    ).fetchone()['count']
                else:
                    total = conn.execute('SELECT COUNT(*) as count FROM avaliacoes').fetchone()['count']
                
                # Contagem por grau de satisfação (com filtro)
                if data_inicio and data_fim:
                    satisfacao_counts = conn.execute('''
                        SELECT grau_satisfacao, COUNT(*) as count 
                        FROM avaliacoes 
                        WHERE data BETWEEN ? AND ?
                        GROUP BY grau_satisfacao
                    ''', (data_inicio, data_fim)).fetchall()
                else:
                    satisfacao_counts = conn.execute('''
                        SELECT grau_satisfacao, COUNT(*) as count 
                        FROM avaliacoes 
                        GROUP BY grau_satisfacao
                    ''').fetchall()
                
                # Contagem por dia da semana
                if data_inicio and data_fim:
                    dia_counts = conn.execute('''
                        SELECT dia_semana, COUNT(*) as count 
                        FROM avaliacoes 
                        WHERE data BETWEEN ? AND ?
                        GROUP BY dia_semana
                        ORDER BY 
                            CASE dia_semana
                                WHEN 'Segunda' THEN 1
                                WHEN 'Terça' THEN 2
                                WHEN 'Quarta' THEN 3
                                WHEN 'Quinta' THEN 4
                                WHEN 'Sexta' THEN 5
                                WHEN 'Sábado' THEN 6
                                WHEN 'Domingo' THEN 7
                            END
                    ''', (data_inicio, data_fim)).fetchall()
                else:
                    dia_counts = conn.execute('''
                        SELECT dia_semana, COUNT(*) as count 
                        FROM avaliacoes 
                        GROUP BY dia_semana
                        ORDER BY 
                            CASE dia_semana
                                WHEN 'Segunda' THEN 1
                                WHEN 'Terça' THEN 2
                                WHEN 'Quarta' THEN 3
                                WHEN 'Quinta' THEN 4
                                WHEN 'Sexta' THEN 5
                                WHEN 'Sábado' THEN 6
                                WHEN 'Domingo' THEN 7
                            END
                    ''').fetchall()
                
                # Avaliações por data (últimos 30 dias ou período filtrado)
                if data_inicio and data_fim:
                    avaliacoes_diarias = conn.execute('''
                        SELECT data, COUNT(*) as count 
                        FROM avaliacoes 
                        WHERE data BETWEEN ? AND ?
                        GROUP BY data 
                        ORDER BY data
                    ''', (data_inicio, data_fim)).fetchall()
                else:
                    avaliacoes_diarias = conn.execute('''
                        SELECT data, COUNT(*) as count 
                        FROM avaliacoes 
                        WHERE date(data) >= date('now', '-30 days')
                        GROUP BY data 
                        ORDER BY data
                    ''').fetchall()
                
                offset = (pagina - 1) * limite
                if data_inicio and data_fim:
                    ultimas_avaliacoes = conn.execute('''
                        SELECT id, grau_satisfacao, data, hora, dia_semana 
                        FROM avaliacoes 
                        WHERE data BETWEEN ? AND ?
                        ORDER BY id DESC 
                        LIMIT ? OFFSET ?
                    ''', (data_inicio, data_fim, limite, offset)).fetchall()
                    
                    total_registos = conn.execute(
                        'SELECT COUNT(*) as count FROM avaliacoes WHERE data BETWEEN ? AND ?',
                        (data_inicio, data_fim)
                    ).fetchone()['count']
                else:
                    ultimas_avaliacoes = conn.execute('''
                        SELECT id, grau_satisfacao, data, hora, dia_semana 
                        FROM avaliacoes 
                        ORDER BY id DESC 
                        LIMIT ? OFFSET ?
                    ''', (limite, offset)).fetchall()
                    
                    total_registos = conn.execute('SELECT COUNT(*) as count FROM avaliacoes').fetchone()['count']

                satisfacao_dict = {row['grau_satisfacao']: row['count'] for row in satisfacao_counts}
                dia_dict = {row['dia_semana']: row['count'] for row in dia_counts}
                avaliacoes_diarias = [{'data': row['data'], 'count': row['count']} for row in avaliacoes_diarias]
                ultimas_avaliacoes = [dict(row) for row in ultimas_avaliacoes]

        total_paginas = (total_registos + limite - 1) // limite

        return jsonify({
            'total': total,
            'satisfacao': satisfacao_dict,
            'dia_semana': dia_dict,
            'avaliacoes_diarias': avaliacoes_diarias,
            'ultimas_avaliacoes': ultimas_avaliacoes,
            'paginacao': {
                'pagina_atual': pagina,
                'total_paginas': total_paginas,
                'total_registos': total_registos,
                'registos_por_pagina': limite
            }
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/health', methods=['GET'])
def health():
    """Endpoint de health check"""
    return jsonify({'status': 'ok'})

@app.route('/api/stats/comparacao', methods=['GET'])
def stats_comparacao():
    """Compara estatísticas entre dois dias"""
    try:
        dia1 = request.args.get('dia1')
        dia2 = request.args.get('dia2')
        
        if not dia1 or not dia2:
            return jsonify({'error': 'Parâmetros dia1 e dia2 obrigatórios'}), 400
        
        if USE_MONGO:
            stats_dia1 = _mongo_contar_por_campo('grau_satisfacao', dia1, dia1)
            stats_dia2 = _mongo_contar_por_campo('grau_satisfacao', dia2, dia2)
            total_dia1 = mongo_collection.count_documents({'data': dia1})
            total_dia2 = mongo_collection.count_documents({'data': dia2})
        else:
            with get_db() as conn:
                # Dia 1
                stats_dia1 = conn.execute('''
                    SELECT grau_satisfacao, COUNT(*) as count 
                    FROM avaliacoes 
                    WHERE data = ?
                    GROUP BY grau_satisfacao
                ''', (dia1,)).fetchall()
                
                total_dia1 = conn.execute(
                    'SELECT COUNT(*) as count FROM avaliacoes WHERE data = ?',
                    (dia1,)
                ).fetchone()['count']
                
                # Dia 2
                stats_dia2 = conn.execute('''
                    SELECT grau_satisfacao, COUNT(*) as count 
                    FROM avaliacoes 
                    WHERE data = ?
                    GROUP BY grau_satisfacao
                ''', (dia2,)).fetchall()
                
                total_dia2 = conn.execute(
                    'SELECT COUNT(*) as count FROM avaliacoes WHERE data = ?',
                    (dia2,)
                ).fetchone()['count']

                stats_dia1 = {row['grau_satisfacao']: row['count'] for row in stats_dia1}
                stats_dia2 = {row['grau_satisfacao']: row['count'] for row in stats_dia2}
        
        return jsonify({
            'dia1': {
                'data': dia1,
                'total': total_dia1,
                'distribuicao': stats_dia1
            },
            'dia2': {
                'data': dia2,
                'total': total_dia2,
                'distribuicao': stats_dia2
            }
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/health', methods=['GET'])

@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    """Exporta dados para Excel (.xlsx)"""
    try:
        if USE_MONGO:
            rows = _mongo_listar_avaliacoes(asc=True)
        else:
            with get_db() as conn:
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT id, grau_satisfacao, data, hora, dia_semana 
                    FROM avaliacoes 
                    ORDER BY id ASC
                ''')
                rows = cursor.fetchall()
        
        if not rows:
            return jsonify({'error': 'Sem dados para exportar'}), 400
        
        if HAS_OPENPYXL:
            # Criar arquivo XLSX com formatação
            wb = Workbook()
            ws = wb.active
            ws.title = "Avaliações"
            
            # Estilos
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Cabeçalho
            headers = ['ID', 'Grau de Satisfação', 'Data', 'Hora', 'Dia da Semana']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            
            # Dados
            emoji_map = {
                'muito_satisfeito': '😊 Muito Satisfeito',
                'satisfeito': '😐 Satisfeito',
                'insatisfeito': '😞 Insatisfeito'
            }
            
            for row_idx, row in enumerate(rows, start=2):
                satisfacao_display = emoji_map.get(row[1], row[1])
                data = [row[0], satisfacao_display, row[2], row[3], row[4]]
                for col, value in enumerate(data, start=1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.value = value
                    cell.border = border
                    if col == 1 or col == 3 or col == 4:
                        cell.alignment = Alignment(horizontal='center')
            
            # Ajustar largura das colunas
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 18
            
            # Salvar em memória
            mem_file = BytesIO()
            wb.save(mem_file)
            mem_file.seek(0)
            
            return send_file(
                mem_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'avaliacoes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        
        else:
            # Fallback para CSV
            output = StringIO()
            writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_ALL)
            writer.writerow(['ID', 'Grau de Satisfação', 'Data', 'Hora', 'Dia da Semana'])
            for row in rows:
                writer.writerow([row[0], row[1], row[2], row[3], row[4]])
            
            output.seek(0)
            mem_file = BytesIO()
            mem_file.write(output.getvalue().encode('utf-8'))
            mem_file.seek(0)
            
            return send_file(
                mem_file,
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'avaliacoes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
            )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export/txt', methods=['POST'])
def export_txt():
    """Exporta dados para TXT com filtro opcional por data"""
    try:
        data = request.get_json() if request.is_json else {}
        data_inicio = data.get('data_inicio')
        data_fim = data.get('data_fim')
        
        if USE_MONGO:
            rows = _mongo_listar_avaliacoes(data_inicio, data_fim, asc=True)
        else:
            with get_db() as conn:
                cursor = conn.cursor()
                
                if data_inicio and data_fim:
                    cursor.execute('''
                        SELECT id, grau_satisfacao, data, hora, dia_semana 
                        FROM avaliacoes 
                        WHERE data BETWEEN ? AND ?
                        ORDER BY id ASC
                    ''', (data_inicio, data_fim))
                else:
                    cursor.execute('''
                        SELECT id, grau_satisfacao, data, hora, dia_semana 
                        FROM avaliacoes 
                        ORDER BY id ASC
                    ''')
                
                rows = cursor.fetchall()
        
        # Criar arquivo TXT em memória
        output = StringIO()
        
        # Cabeçalho
        output.write('=' * 100 + '\n')
        output.write('RELATÓRIO DE AVALIAÇÕES DE SATISFAÇÃO\n')
        output.write('=' * 100 + '\n\n')
        
        if data_inicio and data_fim:
            output.write(f'Período: {data_inicio} até {data_fim}\n')
        else:
            output.write(f'Data de Geração: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}\n')
        
        output.write(f'Total de Registros: {len(rows)}\n')
        output.write('=' * 100 + '\n\n')
        
        # Tabela formatada
        output.write(f"{'ID':<6} | {'Satisfação':<25} | {'Data':<12} | {'Hora':<10} | {'Dia da Semana':<15}\n")
        output.write('-' * 100 + '\n')
        
        # Dados
        for row in rows:
            emoji = '😊' if row[1] == 'muito_satisfeito' else '😐' if row[1] == 'satisfeito' else '😞'
            satisfacao_texto = f"{emoji} {row[1].replace('_', ' ').title()}"
            
            output.write(f"{row[0]:<6} | {satisfacao_texto:<25} | {row[2]:<12} | {row[3]:<10} | {row[4]:<15}\n")
        
        output.write('\n' + '=' * 100 + '\n')
        output.write(f'Fim do Relatório\n')
        output.write('=' * 100 + '\n')
        
        # Converter para bytes
        output.seek(0)
        mem_file = BytesIO()
        mem_file.write(output.getvalue().encode('utf-8'))
        mem_file.seek(0)
        
        return send_file(
            mem_file,
            mimetype='text/plain',
            as_attachment=True,
            download_name=f'avaliacoes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500



if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
